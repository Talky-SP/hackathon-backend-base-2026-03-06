"""
Export tool for generating CSV and Excel files from financial data.

Produces downloadable files with Talky branding for Excel.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Talky brand colors
HEADER_BG_COLOR = "F2764B"  # warm orange (no #)
HEADER_FONT_COLOR = "FFFFFF"  # white

DEFAULT_OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "test_output",
    "task_exports",
)


def generate_csv(
    data: list[dict],
    filename: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
) -> str:
    """Write data to CSV with UTF-8 BOM for Excel compatibility.

    Returns the absolute file path.
    """
    os.makedirs(output_dir, exist_ok=True)
    if not filename.endswith(".csv"):
        filename += ".csv"
    filepath = os.path.join(output_dir, filename)

    if not data:
        Path(filepath).write_text("\ufeff", encoding="utf-8")
        return filepath

    fieldnames = list(data[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)

    logger.info("export_tool | CSV saved: %s (%d rows)", filepath, len(data))
    return filepath


def generate_excel(
    data: dict[str, list[dict]],
    filename: str,
    output_dir: str = DEFAULT_OUTPUT_DIR,
    title: str = "",
) -> str:
    """Write a styled Excel workbook with Talky branding.

    Args:
        data: Mapping of sheet_name -> list of row dicts.
        filename: Output filename (without or with .xlsx).
        output_dir: Directory to write to.
        title: Optional title row at the top of each sheet.

    Returns the absolute file path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, numbers
    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV export")
        # Fallback: export first sheet as CSV
        first_sheet = next(iter(data.values()), [])
        return generate_csv(first_sheet, filename.replace(".xlsx", ".csv"), output_dir)

    os.makedirs(output_dir, exist_ok=True)
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    title_font = Font(bold=True, size=14, color=HEADER_BG_COLOR)
    money_fmt = '#,##0.00'
    int_fmt = '#,##0'

    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars

        start_row = 1
        if title:
            ws.cell(row=1, column=1, value=title).font = title_font
            start_row = 3

        if not rows:
            continue

        # Header row
        fieldnames = list(rows[0].keys())
        for col_idx, col_name in enumerate(fieldnames, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, col_name in enumerate(fieldnames, 1):
                val = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                # Apply number formatting
                if isinstance(val, float):
                    cell.number_format = money_fmt
                elif isinstance(val, int) and not isinstance(val, bool):
                    cell.number_format = int_fmt

        # Auto-column-width
        for col_idx, col_name in enumerate(fieldnames, 1):
            max_len = len(str(col_name))
            for row_data in rows[:100]:  # Sample first 100 rows
                val = row_data.get(col_name, "")
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    wb.save(filepath)
    total_rows = sum(len(rows) for rows in data.values())
    logger.info("export_tool | Excel saved: %s (%d rows, %d sheets)", filepath, total_rows, len(data))
    return filepath


def generate_export(
    data_json: str,
    fmt: str = "xlsx",
    filename: str = "report",
    output_dir: str = DEFAULT_OUTPUT_DIR,
    sheet_config: dict | None = None,
    title: str = "",
) -> dict:
    """High-level export function called by TaskAgent tool handler.

    Args:
        data_json: JSON string — either a list of dicts (single sheet) or
                   a dict of {sheet_name: [rows]} (multi-sheet).
        fmt: "csv" or "xlsx".
        filename: Output filename (extension added automatically).
        output_dir: Where to save.
        sheet_config: Optional config for sheet names when data is a list.
        title: Optional title for Excel sheets.

    Returns dict with success, file_path, filename, format, row_count.
    """
    try:
        parsed = json.loads(data_json) if isinstance(data_json, str) else data_json
    except (json.JSONDecodeError, TypeError) as exc:
        return {"success": False, "error": f"Invalid JSON data: {exc}"}

    # Normalize to multi-sheet format
    if isinstance(parsed, list):
        sheet_name = (sheet_config or {}).get("sheet_name", "Data")
        sheets = {sheet_name: parsed}
    elif isinstance(parsed, dict):
        # Check if it's already {sheet_name: [rows]} or a single row dict
        if all(isinstance(v, list) for v in parsed.values()):
            sheets = parsed
        else:
            sheets = {"Data": [parsed]}
    else:
        return {"success": False, "error": f"Unexpected data type: {type(parsed)}"}

    try:
        if fmt == "csv":
            # CSV only supports single sheet — use first
            first_rows = next(iter(sheets.values()), [])
            filepath = generate_csv(first_rows, filename, output_dir)
        else:
            filepath = generate_excel(sheets, filename, output_dir, title=title)

        total_rows = sum(len(rows) for rows in sheets.values())
        return {
            "success": True,
            "file_path": filepath,
            "filename": os.path.basename(filepath),
            "format": fmt,
            "row_count": total_rows,
            "sheets": list(sheets.keys()),
        }
    except Exception as exc:
        logger.error("export_tool | Failed: %s", exc)
        return {"success": False, "error": str(exc)}
