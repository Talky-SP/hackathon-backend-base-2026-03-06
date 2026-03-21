"""
Excel generation tool — creates formatted Excel workbooks with charts.

Used by the task executor for generating financial reports:
- Cash Flow Forecast (13 weeks)
- Pack Reporting (P&L, Balance, KPIs)
- Modelo 303 (IVA trimestral)
- Aging Analysis
- Client Profitability

Architecture:
- Local: saves to ARTIFACTS_DIR/{task_id}/{filename}
- Production: same local path, then uploaded to S3
"""
from __future__ import annotations

import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

ARTIFACTS_DIR = os.path.join(os.environ.get("TEMP", "/tmp"), "cfo_artifacts")

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NEGATIVE_FONT = Font(color="CC0000")
CURRENCY_FORMAT = '#,##0.00 €'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _ensure_dir(task_id: str) -> str:
    path = os.path.join(ARTIFACTS_DIR, task_id)
    os.makedirs(path, exist_ok=True)
    return path


def _style_header_row(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, max_col: int,
                     currency_cols: list[int] | None = None):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if currency_cols and col in currency_cols:
                cell.number_format = CURRENCY_FORMAT
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = NEGATIVE_FONT


def _style_total_row(ws, row: int, max_col: int, currency_cols: list[int] | None = None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if currency_cols and col in currency_cols:
            cell.number_format = CURRENCY_FORMAT


def _auto_column_widths(ws, min_width: int = 12, max_width: int = 40):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ---------------------------------------------------------------------------
# Public API — generate specific report types
# ---------------------------------------------------------------------------
def generate_table_excel(
    task_id: str,
    filename: str,
    sheets: list[dict],
) -> str:
    """
    Generate a multi-sheet Excel workbook from structured data.

    Each sheet dict:
    {
        "name": "Sheet Name",
        "headers": ["Col A", "Col B", ...],
        "rows": [[val1, val2, ...], ...],
        "currency_cols": [2, 3],        # 1-indexed columns to format as EUR
        "total_row": [None, "TOTAL", 1234.56, ...],  # optional
        "chart": {                       # optional
            "type": "bar|line|pie",
            "title": "Chart Title",
            "data_col": 3,              # 1-indexed column for data
            "label_col": 1,             # 1-indexed column for labels
            "position": "E2",
        }
    }

    Returns the file path.
    """
    dir_path = _ensure_dir(task_id)
    filepath = os.path.join(dir_path, filename)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["name"][:31])
        headers = sheet_def["headers"]
        rows = sheet_def["rows"]
        currency_cols = sheet_def.get("currency_cols", [])

        # Headers
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, 1, len(headers))

        # Data rows
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        _style_data_rows(ws, 2, len(rows) + 1, len(headers), currency_cols)

        # Total row
        total_row = sheet_def.get("total_row")
        if total_row:
            t_row = len(rows) + 2
            for c_idx, val in enumerate(total_row, 1):
                ws.cell(row=t_row, column=c_idx, value=val)
            _style_total_row(ws, t_row, len(headers), currency_cols)

        # Chart
        chart_def = sheet_def.get("chart")
        if chart_def and rows:
            chart_type = chart_def.get("type", "bar")
            data_col = chart_def.get("data_col", 2)
            label_col = chart_def.get("label_col", 1)

            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()

            chart.title = chart_def.get("title", "")
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=data_col, min_row=1, max_row=len(rows) + 1)
            labels_ref = Reference(ws, min_col=label_col, min_row=2, max_row=len(rows) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels_ref)

            ws.add_chart(chart, chart_def.get("position", "E2"))

        _auto_column_widths(ws)

    wb.save(filepath)
    return filepath


def generate_cash_flow_excel(
    task_id: str,
    weeks: list[str],
    inflows: list[float],
    outflows: list[float],
    opening_balance: float = 0.0,
    details: dict | None = None,
) -> str:
    """Generate a 13-week cash flow forecast Excel."""
    dir_path = _ensure_dir(task_id)
    filepath = os.path.join(dir_path, "cash_flow_forecast_13w.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Forecast"

    # Build rows
    net = [inflows[i] - outflows[i] for i in range(len(weeks))]
    cumulative = []
    bal = opening_balance
    for n in net:
        bal += n
        cumulative.append(round(bal, 2))

    headers = ["Concepto"] + weeks
    data_rows = [
        ["Saldo Inicial"] + [opening_balance] + [""] * (len(weeks) - 1),
        ["Cobros (Entradas)"] + [round(v, 2) for v in inflows],
        ["Pagos (Salidas)"] + [round(-v, 2) for v in outflows],
        ["Flujo Neto"] + [round(v, 2) for v in net],
        ["Saldo Acumulado"] + cumulative,
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # Write data
    currency_cols = list(range(2, len(headers) + 1))
    for r_idx, row in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_data_rows(ws, 2, len(data_rows) + 1, len(headers), currency_cols)

    # Style special rows
    for r in [5, 6]:  # Net flow + Cumulative
        _style_total_row(ws, r, len(headers), currency_cols)

    # Line chart for cumulative balance
    chart = LineChart()
    chart.title = "Previsión de Tesorería - 13 Semanas"
    chart.y_axis.title = "EUR"
    chart.x_axis.title = "Semana"
    chart.style = 10
    chart.width = 25
    chart.height = 14

    # Inflows line
    inflows_ref = Reference(ws, min_col=2, max_col=len(weeks) + 1, min_row=3)
    chart.add_data(inflows_ref, from_rows=True, titles_from_data=False)
    chart.series[0].name = "Cobros"

    # Outflows line
    outflows_ref = Reference(ws, min_col=2, max_col=len(weeks) + 1, min_row=4)
    chart.add_data(outflows_ref, from_rows=True, titles_from_data=False)
    chart.series[1].name = "Pagos"

    # Cumulative line
    cum_ref = Reference(ws, min_col=2, max_col=len(weeks) + 1, min_row=6)
    chart.add_data(cum_ref, from_rows=True, titles_from_data=False)
    chart.series[2].name = "Saldo Acumulado"

    labels = Reference(ws, min_col=2, max_col=len(weeks) + 1, min_row=1)
    chart.set_categories(labels)
    ws.add_chart(chart, "A9")

    # Details sheet if provided
    if details:
        for detail_name, detail_data in details.items():
            if not detail_data:
                continue
            ds = wb.create_sheet(title=detail_name[:31])
            if isinstance(detail_data, list) and detail_data:
                if isinstance(detail_data[0], dict):
                    det_headers = list(detail_data[0].keys())
                    for col, h in enumerate(det_headers, 1):
                        ds.cell(row=1, column=col, value=h)
                    _style_header_row(ds, 1, len(det_headers))
                    for r_idx, item in enumerate(detail_data, 2):
                        for c_idx, h in enumerate(det_headers, 1):
                            ds.cell(row=r_idx, column=c_idx, value=item.get(h))
                    _auto_column_widths(ds)

    _auto_column_widths(ws)
    wb.save(filepath)
    return filepath


def generate_modelo_303_excel(task_id: str, data: dict) -> str:
    """Generate Modelo 303 (quarterly VAT) draft Excel."""
    dir_path = _ensure_dir(task_id)
    filepath = os.path.join(dir_path, "modelo_303_borrador.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo 303"

    period = data.get("period", "T1 2026")
    ws.cell(row=1, column=1, value=f"BORRADOR MODELO 303 — {period}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    # IVA Repercutido (ventas)
    row = 3
    ws.cell(row=row, column=1, value="IVA REPERCUTIDO (Ventas)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    headers = ["Tipo IVA", "Base Imponible", "Cuota", "Num. Facturas"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, 4)

    repercutido = data.get("iva_repercutido", [])
    total_base_rep, total_cuota_rep = 0.0, 0.0
    for item in repercutido:
        row += 1
        ws.cell(row=row, column=1, value=item.get("tipo", ""))
        ws.cell(row=row, column=2, value=item.get("base", 0))
        ws.cell(row=row, column=3, value=item.get("cuota", 0))
        ws.cell(row=row, column=4, value=item.get("facturas", 0))
        total_base_rep += item.get("base", 0)
        total_cuota_rep += item.get("cuota", 0)
    _style_data_rows(ws, row - len(repercutido) + 1, row, 4, [2, 3])

    row += 1
    ws.cell(row=row, column=1, value="TOTAL REPERCUTIDO")
    ws.cell(row=row, column=2, value=round(total_base_rep, 2))
    ws.cell(row=row, column=3, value=round(total_cuota_rep, 2))
    _style_total_row(ws, row, 4, [2, 3])

    # IVA Soportado (compras)
    row += 2
    ws.cell(row=row, column=1, value="IVA SOPORTADO (Compras)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, 4)

    soportado = data.get("iva_soportado", [])
    total_base_sop, total_cuota_sop = 0.0, 0.0
    for item in soportado:
        row += 1
        ws.cell(row=row, column=1, value=item.get("tipo", ""))
        ws.cell(row=row, column=2, value=item.get("base", 0))
        ws.cell(row=row, column=3, value=item.get("cuota", 0))
        ws.cell(row=row, column=4, value=item.get("facturas", 0))
        total_base_sop += item.get("base", 0)
        total_cuota_sop += item.get("cuota", 0)
    _style_data_rows(ws, row - len(soportado) + 1, row, 4, [2, 3])

    row += 1
    ws.cell(row=row, column=1, value="TOTAL SOPORTADO DEDUCIBLE")
    ws.cell(row=row, column=2, value=round(total_base_sop, 2))
    ws.cell(row=row, column=3, value=round(total_cuota_sop, 2))
    _style_total_row(ws, row, 4, [2, 3])

    # Resultado
    row += 2
    resultado = round(total_cuota_rep - total_cuota_sop, 2)
    ws.cell(row=row, column=1, value="RESULTADO (Repercutido - Soportado)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=resultado)
    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
    ws.cell(row=row, column=2).font = Font(bold=True, size=14,
                                            color="CC0000" if resultado > 0 else "006600")
    row += 1
    if resultado > 0:
        ws.cell(row=row, column=1, value="→ A INGRESAR en Hacienda")
    else:
        ws.cell(row=row, column=1, value="→ A COMPENSAR / DEVOLVER")
    ws.cell(row=row, column=1).font = Font(italic=True, size=11)

    # Operaciones especiales
    especiales = data.get("operaciones_especiales", {})
    if especiales:
        row += 2
        ws.cell(row=row, column=1, value="OPERACIONES ESPECIALES")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        for key, val in especiales.items():
            row += 1
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT

    _auto_column_widths(ws)
    wb.save(filepath)
    return filepath


def get_artifact_path(task_id: str, filename: str) -> str | None:
    """Get the full path for an artifact file (local filesystem)."""
    path = os.path.join(ARTIFACTS_DIR, task_id, filename)
    return path if os.path.exists(path) else None


def get_artifact_url(task_id: str, filename: str) -> str | None:
    """Get a download URL for an artifact — S3 presigned URL or local API path."""
    from hackathon_backend.services.lambdas.agent.core.storage import get_artifact_url as _s3_url
    return _s3_url(task_id, filename)


def list_artifacts(task_id: str) -> list[dict]:
    """List all artifacts for a task — checks both local and S3."""
    from hackathon_backend.services.lambdas.agent.core.storage import list_artifacts as _s3_list, _use_s3
    if _use_s3():
        return _s3_list(task_id)
    # Local mode
    dir_path = os.path.join(ARTIFACTS_DIR, task_id)
    if not os.path.isdir(dir_path):
        return []
    artifacts = []
    for f in os.listdir(dir_path):
        fp = os.path.join(dir_path, f)
        if os.path.isfile(fp):
            artifacts.append({
                "filename": f,
                "size_bytes": os.path.getsize(fp),
                "path": fp,
                "url": f"/api/tasks/{task_id}/artifacts/{f}",
            })
    return artifacts


def upload_artifact_to_s3(task_id: str, filename: str) -> dict | None:
    """Upload a locally generated artifact to S3 (no-op in local mode)."""
    from hackathon_backend.services.lambdas.agent.core.storage import save_artifact, _use_s3
    if not _use_s3():
        return None
    local_path = os.path.join(ARTIFACTS_DIR, task_id, filename)
    if not os.path.isfile(local_path):
        return None
    with open(local_path, "rb") as f:
        data = f.read()
    return save_artifact(task_id, filename, data)
