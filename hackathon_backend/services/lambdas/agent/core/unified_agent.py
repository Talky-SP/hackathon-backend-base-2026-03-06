"""
Unified Agent v2 — "LLM as Director, Code as Executor"

Single agent loop with tools:
  - dynamo_query: fetch data → returns dataset cards (metadata, not raw items)
  - run_code: execute Python locally on full dataset (analysis + file generation)
  - run_code with existing_files: edit previously generated files locally (fast)

The LLM sees schemas, stats, and samples — NEVER thousands of raw items.
It writes code that runs locally against the full dataset in a sandboxed exec().
"""
from __future__ import annotations

import hashlib
import json
import logging
import os
import sys
import time
import uuid
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Any, Callable

from langfuse import observe

from hackathon_backend.services.lambdas.agent.core.config import (
    traced_completion, is_cancelled, CancelledError,
)
from hackathon_backend.services.lambdas.agent.core.query_agent import (
    _execute_query, _sanitize, _extract_source,
)
from hackathon_backend.services.lambdas.agent.core.code_runner import (
    ARTIFACTS_DIR,
)
from hackathon_backend.services.lambdas.agent.core.data_catalog import (
    get_schema_prompt, get_numeric_fields, get_date_fields,
    get_group_fields, get_slim_fields, ALL_TABLE_NAMES,
)
from hackathon_backend.services.lambdas.agent.core.playbooks import (
    classify_intent, get_playbook_guidance, get_playbook_name, PLAYBOOKS,
)
# Lazy import to avoid circular: subagent_runner imports from unified_agent
def _dispatch_subagents(**kwargs):
    from hackathon_backend.services.lambdas.agent.core.subagent_runner import dispatch_subagents
    return dispatch_subagents(**kwargs)

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Query results cache — reuse data across turns within the same chat
# ---------------------------------------------------------------------------
_QUERY_CACHE: dict[str, dict] = {}
_QUERY_CACHE_TTL = 600  # 10 minutes


def _cache_query_results(chat_id: str, query_results: dict[str, dict]) -> None:
    if not chat_id or not query_results:
        return
    # Only cache queries that returned data (skip 0-item results)
    useful = {k: v for k, v in query_results.items()
              if v.get("count", len(v.get("items", []))) > 0}
    if not useful:
        return
    _QUERY_CACHE[chat_id] = {"results": useful, "ts": time.time()}
    if len(_QUERY_CACHE) > 50:
        oldest = sorted(_QUERY_CACHE, key=lambda k: _QUERY_CACHE[k]["ts"])
        for k in oldest[:len(_QUERY_CACHE) - 50]:
            del _QUERY_CACHE[k]


def _get_cached_query_results(chat_id: str) -> dict[str, dict] | None:
    if not chat_id or chat_id not in _QUERY_CACHE:
        return None
    entry = _QUERY_CACHE[chat_id]
    if time.time() - entry["ts"] > _QUERY_CACHE_TTL:
        del _QUERY_CACHE[chat_id]
        return None
    return entry["results"]


# ---------------------------------------------------------------------------
# Multimodal
# ---------------------------------------------------------------------------
_IMAGE_MIMES = {"image/png", "image/jpeg", "image/gif", "image/webp"}
_PDF_MIMES = {"application/pdf"}

EventCallback = Callable[[str, dict], None]


def _noop(event: str, data: dict) -> None:
    pass


# ---------------------------------------------------------------------------
# Context compression — shrink old tool results to save tokens
# ---------------------------------------------------------------------------
_COMPRESS_AFTER_ITERATION = 3  # Start compressing after this many iterations
_KEEP_RECENT_TOOL_RESULTS = 4  # Keep the N most recent tool results in full


def _compress_messages(messages: list[dict]) -> list[dict]:
    """Compress older tool result messages to short summaries.

    Keeps the system prompt, user messages, assistant messages, and the
    most recent tool results untouched. Older tool results are replaced
    with a short summary (table name, item count, success/error).
    """
    # Find all tool result indices
    tool_indices = [
        i for i, m in enumerate(messages) if m.get("role") == "tool"
    ]
    if len(tool_indices) <= _KEEP_RECENT_TOOL_RESULTS:
        return messages  # Nothing to compress

    # Indices to compress (all except the most recent N)
    to_compress = set(tool_indices[:-_KEEP_RECENT_TOOL_RESULTS])

    compressed = []
    for i, msg in enumerate(messages):
        if i in to_compress:
            compressed.append(_summarize_tool_result(msg))
        else:
            compressed.append(msg)
    return compressed


def _summarize_tool_result(msg: dict) -> dict:
    """Replace a tool result message content with a short summary."""
    content = msg.get("content", "")
    try:
        data = json.loads(content) if isinstance(content, str) else content
    except (json.JSONDecodeError, TypeError):
        # If content is very long text, truncate it
        summary = content[:200] + "..." if len(content) > 200 else content
        return {**msg, "content": summary}

    if isinstance(data, dict):
        # Dataset card summary
        if "total_items" in data:
            summary = {
                "compressed": True,
                "table": data.get("table", "?"),
                "total_items": data.get("total_items", 0),
                "stats_keys": list(data.get("stats", {}).keys()),
            }
        # Code execution summary
        elif "success" in data:
            summary = {
                "compressed": True,
                "success": data.get("success"),
                "error": data.get("error"),
                "files": [f.get("filename", "?") for f in data.get("files", [])],
                "result_preview": str(data.get("result", ""))[:200] if data.get("result") else None,
            }
        else:
            summary = {"compressed": True, "keys": list(data.keys())[:10]}
        return {**msg, "content": json.dumps(summary)}

    return {**msg, "content": str(data)[:300]}


# ---------------------------------------------------------------------------
# Pre-aggregation helpers (injected into run_code sandbox)
# ---------------------------------------------------------------------------
def _group_by(items, field, agg_field="total", agg_fn="sum"):
    """group_by(items, 'category', 'total') -> {'Alimentacion': 15000.0, ...}"""
    groups: dict[str, list] = {}
    for it in items:
        key = str(it.get(field, "Unknown"))
        val = float(it.get(agg_field, 0) or 0)
        groups.setdefault(key, []).append(val)
    fns = {
        "sum": sum, "count": len,
        "avg": lambda v: sum(v) / len(v) if v else 0,
        "min": min, "max": max,
    }
    fn = fns.get(agg_fn, sum)
    return {k: round(fn(v), 2) for k, v in groups.items()}


def _monthly_totals(items, date_field="invoice_date", amount_field="total"):
    """monthly_totals(items, 'pnl_date') -> {'2026-01': 12000.0, ...}"""
    totals: dict[str, float] = {}
    for it in items:
        month = str(it.get(date_field, ""))[:7]
        if month:
            totals[month] = totals.get(month, 0) + float(it.get(amount_field, 0) or 0)
    return {k: round(v, 2) for k, v in sorted(totals.items())}


def _top_n(items, field, n=10, sort_field="total", reverse=True):
    """top_n(items, 'supplier', 5) -> [{name, total, count}, ...]"""
    groups: dict[str, dict] = {}
    for it in items:
        key = it.get(field)
        if key:
            g = groups.setdefault(str(key), {"key": str(key), "sum": 0, "count": 0})
            g["sum"] += float(it.get(sort_field, 0) or 0)
            g["count"] += 1
    ranked = sorted(groups.values(), key=lambda x: x["sum"], reverse=reverse)
    return [{"name": g["key"], "total": round(g["sum"], 2), "count": g["count"]}
            for g in ranked[:n]]


def _filter_items(items, **conditions):
    """filter_items(items, status='PENDING', reconciled=False) -> filtered list"""
    result = []
    for it in items:
        if all(it.get(k) == v for k, v in conditions.items()):
            result.append(it)
    return result


def _sum_field(items, field):
    """sum_field(items, 'total') -> 145230.50"""
    return round(sum(float(it.get(field, 0) or 0) for it in items), 2)


# ---------------------------------------------------------------------------
# Sandboxed code execution (_safe_exec)
# ---------------------------------------------------------------------------
_SAFE_BUILTINS = {
    # Data types
    "int": int, "float": float, "str": str, "bool": bool,
    "list": list, "dict": dict, "set": set, "tuple": tuple,
    "bytes": bytes, "bytearray": bytearray, "type": type,
    # Iteration
    "range": range, "enumerate": enumerate, "zip": zip,
    "map": map, "filter": filter, "reversed": reversed,
    # Aggregation
    "len": len, "sum": sum, "min": min, "max": max, "abs": abs,
    "round": round, "sorted": sorted, "any": any, "all": all,
    # Utilities
    "isinstance": isinstance, "hasattr": hasattr, "getattr": getattr,
    "print": print, "repr": repr, "format": format,
    "ValueError": ValueError, "TypeError": TypeError, "KeyError": KeyError,
    "IndexError": IndexError, "Exception": Exception, "StopIteration": StopIteration,
    "True": True, "False": False, "None": None,
    # NO: open, __import__, eval, exec, compile, globals, locals, dir,
    #     vars, setattr, delattr, input, breakpoint, exit, quit
}


def _safe_exec(code: str, query_results: dict, file_task_id: str,
               existing_artifacts: list[dict] | None = None) -> dict:
    """Execute code in a sandboxed environment with data injection.

    existing_artifacts: list of {"filename": str, "path": str, "task_id": str}
        from previous turns — accessible via existing_files dict in the sandbox.
    """
    output_dir = os.path.join(ARTIFACTS_DIR, file_task_id)
    os.makedirs(output_dir, exist_ok=True)
    log.info(f"[_safe_exec] task={file_task_id}, output_dir={output_dir}, queries={list(query_results.keys())}")

    # Check memory before execution
    data_json_str = json.dumps(query_results, default=str)
    data_size_mb = len(data_json_str) / (1024 * 1024)
    log.info(f"[_safe_exec] data_size={data_size_mb:.2f}MB, code_len={len(code)}")
    if data_size_mb > 500:
        return {
            "success": False, "result": None, "files": [],
            "error": f"Dataset too large ({data_size_mb:.0f}MB). Use more specific queries.",
        }

    # Pre-import libraries
    injected = {}
    try:
        import pandas
        injected["pd"] = pandas
    except ImportError:
        pass
    try:
        import openpyxl
        import openpyxl.styles
        import openpyxl.utils
        import openpyxl.chart
        import openpyxl.chart.label
        import openpyxl.formatting
        import openpyxl.formatting.rule
        injected["openpyxl"] = openpyxl
    except ImportError:
        pass
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        injected["plt"] = plt
        injected["matplotlib"] = matplotlib
    except ImportError:
        pass
    try:
        import numpy
        injected["np"] = numpy
    except ImportError:
        pass

    # Restricted open() — write only in output_dir, read also from any artifact dir
    _real_open = open
    _allowed_write_dir = os.path.normpath(output_dir)
    _artifacts_root = os.path.normpath(ARTIFACTS_DIR)

    def _sandbox_open(path, mode="r", *args, **kwargs):
        norm = os.path.normpath(str(path))
        is_read = not any(m in mode for m in ("w", "a", "x", "+"))
        if is_read and norm.startswith(_artifacts_root):
            # Allow reading any existing artifact (for edit workflows)
            return _real_open(path, mode, *args, **kwargs)
        if not norm.startswith(_allowed_write_dir):
            raise PermissionError(f"Cannot open '{path}' — use output_dir for file operations")
        return _real_open(path, mode, *args, **kwargs)

    # Safe __import__ — only allow pre-loaded modules (no arbitrary imports)
    _allowed_modules = set(injected.keys()) | {
        "json", "datetime", "collections", "decimal", "math", "re", "statistics",
        "time", "_strptime",  # needed internally by datetime.strptime()
        "openpyxl", "openpyxl.styles", "openpyxl.utils", "openpyxl.chart",
        "openpyxl.chart.label", "openpyxl.chart.series",
        "openpyxl.formatting", "openpyxl.formatting.rule",
        "numpy", "matplotlib", "matplotlib.pyplot",
    }

    def _safe_import(name, globals=None, locals=None, fromlist=(), level=0):
        if level != 0:
            raise ImportError(f"Relative imports not allowed")
        if name not in _allowed_modules:
            # Check if it's a submodule of an allowed package
            parts = name.split(".")
            if not any(name.startswith(m + ".") or name == m for m in _allowed_modules):
                raise ImportError(f"Import of '{name}' is not allowed. Use pre-injected libraries.")
        return __builtins__["__import__"](name, globals, locals, fromlist, level) if isinstance(__builtins__, dict) else __import__(name, globals, locals, fromlist, level)

    builtins_with_open = {**_SAFE_BUILTINS, "open": _sandbox_open, "__import__": _safe_import}

    # Build existing_files map: {filename: absolute_path} for artifacts from previous turns
    _existing_files: dict[str, str] = {}
    for art in (existing_artifacts or []):
        art_path = art.get("path", "")
        if art_path and os.path.isfile(art_path):
            _existing_files[art["filename"]] = art_path
        else:
            # Try to find in artifacts dir by task_id
            candidate = os.path.join(ARTIFACTS_DIR, art.get("task_id", ""), art.get("filename", ""))
            if os.path.isfile(candidate):
                _existing_files[art["filename"]] = candidate

    safe_globals = {
        "__builtins__": builtins_with_open,
        # Data
        "data": query_results,
        "output_dir": output_dir,
        "existing_files": _existing_files,  # {filename: path} — read-only access to previous artifacts
        "result": None,
        # Standard lib
        "json": json, "Decimal": Decimal,
        "datetime": datetime, "timedelta": timedelta,
        "Counter": Counter, "defaultdict": defaultdict,
        # Libraries
        **injected,
        # Helpers
        "group_by": _group_by, "monthly_totals": _monthly_totals,
        "top_n": _top_n, "filter_items": _filter_items, "sum_field": _sum_field,
    }

    # Rewrite /tmp/ paths to output_dir (LLM often writes to /tmp/)
    normalized_output = output_dir.replace("\\", "/")
    code = code.replace("/tmp/", normalized_output + "/")

    t0 = time.time()
    try:
        exec(code, safe_globals)
        elapsed_ms = int((time.time() - t0) * 1000)
    except Exception as e:
        elapsed_ms = int((time.time() - t0) * 1000)
        log.error(f"[_safe_exec] FAILED after {elapsed_ms}ms: {type(e).__name__}: {e}")
        print(f"[_safe_exec] FAILED: {type(e).__name__}: {e}", flush=True)
        return {
            "success": False, "result": None, "files": [],
            "error": f"{type(e).__name__}: {e}", "elapsed_ms": elapsed_ms,
        }

    log.info(f"[_safe_exec] exec OK in {elapsed_ms}ms, scanning output_dir={output_dir}")
    result_val = safe_globals.get("result")
    files = []
    if os.path.isdir(output_dir):
        all_entries = os.listdir(output_dir)
        log.info(f"[_safe_exec] output_dir contents: {all_entries}")
        for f in all_entries:
            fp = os.path.join(output_dir, f)
            if os.path.isfile(fp) and not f.startswith("_"):
                fsize = os.path.getsize(fp)
                log.info(f"[_safe_exec] Found file: {f} ({fsize} bytes)")
                files.append({
                    "filename": f, "path": fp,
                    "size_bytes": fsize,
                    "type": _detect_file_type(f),
                })
    else:
        log.warning(f"[_safe_exec] output_dir does not exist: {output_dir}")

    if not files and not result_val:
        log.warning("[_safe_exec] No files and no result produced by code execution")

    return {
        "success": True,
        "result": _sanitize(result_val) if result_val else None,
        "files": files,
        "error": None,
        "elapsed_ms": elapsed_ms,
    }


def _detect_file_type(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return {
        "xlsx": "excel", "xls": "excel", "csv": "csv",
        "pdf": "pdf", "png": "image", "jpg": "image", "jpeg": "image",
        "json": "json", "html": "html",
    }.get(ext, "file")


# ---------------------------------------------------------------------------
# Audit trail for code execution
# ---------------------------------------------------------------------------
def _audit_code_execution(
    task_id: str, code: str, query_results: dict,
    result: Any, error: str | None, elapsed_ms: int,
    files: list[dict] | None = None,
    location_id: str = "", chat_id: str | None = None,
) -> None:
    """Immutable audit record of every code execution."""
    data_summary = {
        k: {"count": len(v.get("items", [])), "table": v.get("table", "")}
        for k, v in query_results.items()
    }
    data_hash = hashlib.sha256(
        json.dumps(data_summary, sort_keys=True).encode()
    ).hexdigest()[:16]

    try:
        from hackathon_backend.services.lambdas.agent.core.chat_store import record_trace
        record_trace(
            step="code_execution_audit",
            location_id=location_id,
            task_id=task_id,
            chat_id=chat_id,
            model="local_exec",
            provider="local",
            input_data={
                "code": code[:10000],
                "data_hash": data_hash,
                "data_summary": data_summary,
            },
            output_data={
                "success": error is None,
                "error": error[:1000] if error else None,
                "result_type": type(result).__name__ if result else None,
                "result_preview": str(result)[:500] if result else None,
                "files": [f["filename"] for f in (files or [])],
            },
            prompt_tokens=0, completion_tokens=0, total_tokens=0,
            latency_ms=elapsed_ms,
            status="ok" if error is None else "error",
        )
    except Exception as e:
        log.warning(f"[audit] Failed to record code execution audit: {e}")


# ---------------------------------------------------------------------------
# Output validation for generated files
# ---------------------------------------------------------------------------
def _validate_generated_files(files: list[dict]) -> list[dict]:
    """Validate generated files and add metrics."""
    validated = []
    for f in files:
        info = {**f, "valid": True, "metrics": {}}
        fp = f["path"]
        if fp.endswith(".xlsx") or fp.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(fp, read_only=True)
                sheets = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    rows = ws.max_row or 0
                    cols = ws.max_column or 0
                    sheets[name] = {"rows": rows, "cols": cols}
                wb.close()
                info["metrics"] = {
                    "sheets": sheets,
                    "total_rows": sum(s["rows"] for s in sheets.values()),
                }
                if all(s["rows"] <= 1 for s in sheets.values()):
                    info["valid"] = False
                    info["warning"] = "Excel has no data rows (only headers)"
            except Exception as e:
                info["valid"] = False
                info["warning"] = f"Cannot validate: {e}"
        elif fp.endswith(".csv"):
            try:
                with open(fp, "r") as fh:
                    lines = sum(1 for _ in fh)
                info["metrics"] = {"lines": lines}
                if lines <= 1:
                    info["valid"] = False
                    info["warning"] = "CSV has no data rows"
            except Exception:
                pass
        validated.append(info)
    return validated


# ---------------------------------------------------------------------------
# Dataset cards — replace raw items in LLM context
# ---------------------------------------------------------------------------
def _build_dataset_card(query_key: str, result: dict, table_name: str) -> dict:
    """Build a dataset card with stats over ALL items (fully paginated)."""
    items = result.get("items", [])
    total_count = len(items)

    card: dict[str, Any] = {
        "query_key": query_key,
        "table": table_name,
        "total_items": total_count,
        "fields": list(items[0].keys()) if items else [],
    }

    if total_count == 0:
        card["access"] = f"No items found. data['{query_key}']['items'] is empty."
        return card

    # Numeric stats over ALL items
    for field in get_numeric_fields(table_name):
        values = []
        for it in items:
            v = it.get(field)
            if v is not None and v != "":
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    pass
        if values:
            card.setdefault("stats", {})[field] = {
                "sum": round(sum(values), 2),
                "min": round(min(values), 2),
                "max": round(max(values), 2),
                "avg": round(sum(values) / len(values), 2),
                "count_non_null": len(values),
            }

    # Payroll special handling — extract nested numeric fields
    if table_name == "Payroll_Slips" and items:
        for pf in ["gross_amount", "net_amount", "company_ss_contribution", "irpf_amount"]:
            values = []
            for it in items:
                pi = it.get("payroll_info") or {}
                v = pi.get(pf)
                if v is not None:
                    try:
                        values.append(float(v))
                    except (ValueError, TypeError):
                        pass
            if values:
                card.setdefault("stats", {})[f"payroll_info.{pf}"] = {
                    "sum": round(sum(values), 2),
                    "min": round(min(values), 2),
                    "max": round(max(values), 2),
                    "avg": round(sum(values) / len(values), 2),
                    "count_non_null": len(values),
                }

    # Date range over ALL items
    for df in get_date_fields(table_name):
        dates = sorted([str(it[df]) for it in items if it.get(df)])
        if dates:
            card["date_range"] = {"field": df, "min": dates[0], "max": dates[-1]}
            break

    # Distributions over ALL items (top 5)
    for gf in get_group_fields(table_name):
        groups: dict[str, int] = {}
        for it in items:
            k = it.get(gf)
            if k:
                groups[str(k)] = groups.get(str(k), 0) + 1
        if groups:
            top = sorted(groups.items(), key=lambda x: -x[1])[:5]
            remaining = total_count - sum(v for _, v in top)
            dist = dict(top)
            if remaining > 0:
                dist["_other"] = remaining
            card.setdefault("distributions", {})[gf] = dist

    # 3 sample rows (slim fields for pattern recognition)
    slim = get_slim_fields(table_name)
    if slim:
        card["sample_rows"] = [
            {k: v for k, v in it.items() if k in slim}
            for it in items[:3]
        ]
    else:
        card["sample_rows"] = [
            {k: v for k, v in it.items()}
            for it in items[:3]
        ]

    card["access"] = (
        f"All {total_count} items stored in data['{query_key}']['items']. "
        f"Use run_code to process them."
    )
    return card


# ---------------------------------------------------------------------------
# System prompt — uses data_catalog for schema
# ---------------------------------------------------------------------------
def _build_system_prompt(extra_system: str = "", location_id: str = "") -> list[dict]:
    """Build system prompt as cached content blocks."""
    schema_text = get_schema_prompt()

    rules_text = """\
You are an expert AI CFO assistant (Controller Financiero IA). You help business
owners understand their financial data in real time.

TOOLS:
- `dynamo_query`: Query DynamoDB tables. locationId is auto-enforced.
  Returns a DATASET CARD (metadata: stats, distributions, sample rows) — NOT raw items.
  All items are stored in memory for run_code to access.
- `run_code`: Execute Python code with full access to ALL queried data.
  Use for: analysis, aggregation, charts, AND file generation (Excel, CSV, PDF).
  Environment: data dict, output_dir, openpyxl, matplotlib (plt), numpy (np). NO pandas — use basic Python.
  Helpers: group_by(), monthly_totals(), top_n(), filter_items(), sum_field().
- `dispatch_subagents`: Fan out work to parallel AI subagents. Each subagent gets documents
  + full tool access (vision, queries, code execution). Use when processing 5+ documents.
  Each subagent extracts structured data. You consolidate the results.

MULTIMODAL INPUT:
- Users can attach images (PNG, JPG) and PDFs — you can see and analyze them.
- For MANY documents (5+), use dispatch_subagents to process them in parallel batches.

WORKFLOW:
1. Simple questions: Answer directly.
2. Data questions: dynamo_query -> read dataset card stats -> run_code for detailed analysis.
3. File generation: dynamo_query -> run_code (write to output_dir).
4. Edit existing file: run_code with existing_files dict to load, modify, and save.
5. Batch document processing (5+ documents):
   a. Split documents into batches of 3-5 per subagent.
   b. Call dispatch_subagents with objective + document paths.
   c. Subagents process docs in parallel (OCR, extract data, cross-ref DB).
   d. You receive consolidated structured results from all subagents.
   e. Use run_code to merge results, generate Excel, create journal entries, etc.

DATASET CARDS:
- dynamo_query returns metadata (field stats, distributions, 3 sample rows).
- To access actual items, use run_code: `items = data['query_1']['items']`
- Stats in the card are computed over ALL items (not a sample).

run_code ENVIRONMENT:
- `data`: dict of all query results. `data['query_1']['items']` = list of dicts.
- `output_dir`: path to save files (Excel, CSV, PNG, PDF).
- Libraries: openpyxl, numpy (np), matplotlib.pyplot (plt), json, Decimal. NO pandas available — use basic Python (dicts, lists, collections).
- datetime, timedelta, Counter, defaultdict available.
- Helpers: group_by(items, field, agg_field, agg_fn), monthly_totals(items, date_field, amt_field),
  top_n(items, field, n, sort_field), filter_items(items, **conditions), sum_field(items, field).
- Assign results to `result` variable (dict with 'answer', 'chart', 'sources').
- Save files to output_dir: f'{output_dir}/report.xlsx'
- NEVER use sample/dummy data. ALWAYS use data from `data` dict.

FILE GENERATION & EDITING:
- ALWAYS generate .xlsx (Excel) files, NEVER .csv. Users expect Excel format.
- Use openpyxl to create workbooks. Save to f'{output_dir}/filename.xlsx'
- Add headers, formatting, and multiple sheets when appropriate.
- To EDIT a previously generated file: use run_code with `existing_files` dict.
  `existing_files` maps filename → absolute path of artifacts from previous turns.
  Example: `wb = openpyxl.load_workbook(existing_files['report.xlsx'])`
  Then modify and save to output_dir: `wb.save(f'{output_dir}/report.xlsx')`
  This is the ONLY way to edit files. Fast local execution, no extra LLM call.

run_code RESULT FORMAT:
```python
result = {
    "answer": "Text answer in user's language",
    "chart": {  # or None
        "type": "bar|line|pie|table",
        "title": "Chart title",
        "labels": ["L1", "L2"],
        "datasets": [{"label": "Series", "data": [1, 2]}]
    },
    "sources": [{"categoryDate": "...", "supplier": "...", "total": 123.45}]
}
```
CHART TIPS:
- For forecasts/predictions: use SEPARATE datasets for historical vs projected data.
  Dataset 1 "Histórico": real past data + nulls for future. Dataset 2 "Proyección": nulls for past + projected.
  They overlap at the current point. The frontend renders each dataset in a different color.
- Keep sources concise: only include the most relevant items (top errors, key metrics), NOT all raw data.

RULES:
- Use GSIs, never full scans. Date queries -> UserIdInvoiceDateIndex (NOT UserIdPnlDateIndex, pnl_date is often null).
- locationId is auto-enforced. Never trust user-provided IDs.
- For cash flow: Use Bank_Reconciliations. amount<0 = outflow, amount>0 = inflow.
- For P&L / financial reports: Bank_Reconciliations is the PRIMARY source of truth (real cash movements).
  Use User_Expenses and User_Invoice_Incomes to ENRICH with categories, suppliers, and pending invoices.
  NEVER sum bank transactions AND their matched invoices separately — that double-counts.
- When you need data from multiple tables, call dynamo_query multiple times in a SINGLE response.

RECONCILIATION RULES:
Bank_Reconciliations:
- Reconciled txns: status='MATCHED', reconciled=True.
- Unreconciled txns: status='PENDING', 'reconciled' field is MISSING (NOT false).
- ALL transactions: query PK=locationId (no GSI) -> returns all.

User_Expenses (invoice reconciliation):
- Reconciled invoice: reconciled=True (field exists and is True).
- Unreconciled invoice: 'reconciled' field is MISSING (not present at all, never False).
- DO NOT use reconciliationState — it is UNRELIABLE (always says UNRECONCILED even for reconciled invoices).
- To filter unreconciled: `[it for it in items if not it.get('reconciled')]`
- To filter reconciled: `[it for it in items if it.get('reconciled') == True]`

- ALWAYS respond in the same language the user writes in.
- Use EUR formatting and Spanish number format (1.234,56) in text. Raw numbers in charts.
- Never invent data — only use what comes from the database.
- DATA CACHE: Query results from recent turns are cached. If user asks to work with
  recently queried data, run_code can access it directly without re-querying.

TODAY'S DATE: 2026-03-22."""

    if extra_system:
        rules_text += f"\n\nADDITIONAL CONTEXT:\n{extra_system}"
    rules_text += f"\nCURRENT CONTEXT: locationId={location_id}"

    return [
        {"type": "text", "text": schema_text, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": rules_text, "cache_control": {"type": "ephemeral"}},
    ]


# ---------------------------------------------------------------------------
# Multimodal user message builder
# ---------------------------------------------------------------------------
def _build_user_content(
    text: str, attachments: list[dict] | None = None,
) -> str | list[dict]:
    if not attachments:
        return text

    # If many documents (5+), hint subagent dispatch and include paths
    saved_paths = [a for a in attachments if a.get("saved_path")]
    if len(saved_paths) >= 5:
        path_list = "\n".join(f"  - {a.get('filename', '?')}: {a['saved_path']}" for a in saved_paths)
        dispatch_hint = (
            f"\n\n[BATCH DOCUMENTS: {len(saved_paths)} files attached. "
            f"Use dispatch_subagents to process them in parallel. "
            f"Split into batches of 3-5 docs per subagent. File paths:\n{path_list}]"
        )
        text = text + dispatch_hint

    blocks: list[dict] = [{"type": "text", "text": text}]

    # For small batches (< 5 docs), send inline as multimodal
    # For large batches (5+), only send first 2 as preview + use dispatch for the rest
    inline_limit = 4 if len(attachments) < 5 else 2
    for i, att in enumerate(attachments):
        mime = att.get("mime_type", "application/octet-stream")
        b64 = att.get("data", "")
        fname = att.get("filename", "file")

        if i >= inline_limit and len(attachments) >= 5:
            # Skip inline — these will be processed by subagents
            if i == inline_limit:
                blocks.append({"type": "text", "text": f"[... and {len(attachments) - inline_limit} more documents — use dispatch_subagents with saved_path to process them]"})
            continue

        if mime in _IMAGE_MIMES:
            blocks.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
        elif mime in _PDF_MIMES:
            blocks.append({"type": "file", "file": {"file_data": f"data:{mime};base64,{b64}"}})
        else:
            blocks.append({"type": "text", "text": f"[Attached file: {fname} ({mime}) — unsupported]"})
    return blocks


def _build_artifact_context(chat_artifacts: list[dict] | None) -> str:
    if not chat_artifacts:
        return ""
    lines = ["PREVIOUSLY GENERATED FILES (available in existing_files dict in run_code):"]
    for a in chat_artifacts:
        fname = a.get("filename", "?")
        tid = a.get("task_id", "?")
        url = a.get("url", "")
        lines.append(f"  - {fname} (task_id={tid}, url={url})")
    lines.append(
        "\nIMPORTANT: To EDIT these files you MUST call run_code. "
        "Do NOT just say you did it — you must actually execute code. "
        "Load via: wb = openpyxl.load_workbook(existing_files['filename.xlsx'])"
    )
    return "\n".join(lines)


def _build_state_summary(
    cached_queries: dict[str, dict] | None,
    chat_artifacts: list[dict] | None,
) -> str:
    """Build a concise state summary injected before the user message.

    This tells the LLM what data and files are already available so it
    doesn't re-query or hallucinate actions.
    """
    parts: list[str] = []

    if cached_queries:
        lines = ["[CACHED DATA — already in memory, no need to re-query]:"]
        for qk, qv in cached_queries.items():
            table = qv.get("table", "?")
            count = qv.get("count", len(qv.get("items", [])))
            lines.append(f"  - data['{qk}']: {table} ({count} items)")
        lines.append("Use run_code to access this data directly. Only call dynamo_query if you need DIFFERENT data.")
        parts.append("\n".join(lines))

    if chat_artifacts:
        lines = ["[EXISTING FILES — available in existing_files dict in run_code]:"]
        for a in chat_artifacts:
            fname = a.get("filename", "?")
            lines.append(f"  - existing_files['{fname}']")
        lines.append("To EDIT: call run_code, load with openpyxl.load_workbook(existing_files['name.xlsx']), modify, save to output_dir.")
        parts.append("\n".join(lines))

    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Tool definitions
# ---------------------------------------------------------------------------
UNIFIED_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "dynamo_query",
            "description": "Execute a DynamoDB query. locationId is auto-enforced. Returns a dataset card (stats, distributions, samples) — not raw items. Full data accessible via run_code.",
            "parameters": {
                "type": "object",
                "properties": {
                    "table_name": {
                        "type": "string",
                        "description": "Table name without stage prefix",
                        "enum": ALL_TABLE_NAMES,
                    },
                    "index_name": {
                        "type": "string",
                        "description": "GSI name (null for primary key query)",
                    },
                    "pk_field": {
                        "type": "string",
                        "description": "Partition key field name (default: userId)",
                        "default": "userId",
                    },
                    "pk_value": {
                        "type": "string",
                        "description": "PK value. For userId/locationId this is auto-set. Only provide for composite keys.",
                    },
                    "sk_field": {
                        "type": "string",
                        "description": "Sort key field name",
                    },
                    "sk_condition": {
                        "type": "object",
                        "description": "Sort key condition",
                        "properties": {
                            "op": {"type": "string", "enum": ["eq", "between", "begins_with", "gt", "lt"]},
                            "value": {"type": "string"},
                            "value2": {"type": "string", "description": "Second value for 'between'"},
                        },
                        "required": ["op", "value"],
                    },
                    "filter_expression": {
                        "description": "Post-query filter(s). Single object or array.",
                        "oneOf": [
                            {
                                "type": "object",
                                "properties": {
                                    "field": {"type": "string"},
                                    "op": {"type": "string", "enum": ["eq", "ne", "contains", "begins_with", "exists", "gt", "lt"]},
                                    "value": {},
                                },
                                "required": ["field", "op", "value"],
                            },
                            {"type": "array", "items": {"type": "object"}},
                        ],
                    },
                    "limit": {"type": "integer", "description": "Max items to return"},
                },
                "required": ["table_name"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_code",
            "description": (
                "Execute Python code with full access to ALL queried data. "
                "Use for analysis, aggregation, chart generation, and file creation. "
                "\n\nENVIRONMENT:\n"
                "- `data`: dict of all query results. data['query_1']['items'] = list of dicts.\n"
                "- `output_dir`: path to save files (Excel, CSV, PNG, PDF).\n"
                "- `existing_files`: dict mapping filename→path for previously generated artifacts.\n"
                "  To edit an existing file: wb = openpyxl.load_workbook(existing_files['name.xlsx'])\n"
                "- Libraries: openpyxl, numpy (np), matplotlib.pyplot (plt), json, Decimal. NO pandas available — use basic Python (dicts, lists, collections).\n"
                "- datetime, timedelta, Counter, defaultdict.\n"
                "- Helpers: group_by(items, field, agg_field='total', agg_fn='sum'), "
                "monthly_totals(items, date_field, amount_field), "
                "top_n(items, field, n=10, sort_field='total'), "
                "filter_items(items, **conditions), sum_field(items, field).\n"
                "\n\nRULES:\n"
                "- Assign analysis results to `result` variable (dict with 'answer', 'chart', 'sources').\n"
                "- Save generated files to output_dir (e.g. f'{output_dir}/report.xlsx').\n"
                "- NEVER use sample/dummy data. Always use data from `data` dict.\n"
                "- Code runs in a sandboxed environment. No imports needed — all libraries pre-injected."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Python code to execute.",
                    },
                },
                "required": ["code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "dispatch_subagents",
            "description": (
                "Fan out work to parallel subagents for batch document processing. "
                "Each subagent receives a batch of documents and an objective, and has "
                "full access to vision (multimodal), DynamoDB queries, and code execution. "
                "\n\nUSE WHEN:\n"
                "- User uploads many documents (invoices, receipts, payslips) that need individual processing.\n"
                "- A task can be split into independent parallel subtasks.\n"
                "- Processing 5+ documents that each need OCR/analysis.\n"
                "\n\nEach subagent extracts structured data from its documents, "
                "can cross-reference with the database, and returns results. "
                "You (the main agent) then consolidate all subagent results into the final output."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "subtasks": {
                        "type": "array",
                        "description": "List of subtasks. Each gets its own parallel subagent.",
                        "items": {
                            "type": "object",
                            "properties": {
                                "objective": {
                                    "type": "string",
                                    "description": "What this subagent should do with its documents.",
                                },
                                "documents": {
                                    "type": "array",
                                    "items": {"type": "string"},
                                    "description": "File paths of documents to process.",
                                },
                            },
                            "required": ["objective", "documents"],
                        },
                    },
                    "max_parallel": {
                        "type": "integer",
                        "description": "Max subagents running simultaneously (default 5, max 10).",
                        "default": 5,
                    },
                },
                "required": ["subtasks"],
            },
        },
    },
]


# ---------------------------------------------------------------------------
# Main agent loop
# ---------------------------------------------------------------------------
@observe(name="unified_agent")
def run_agent(
    user_message: str,
    location_id: str,
    model_id: str = "claude-sonnet-4.5",
    conversation_history: list[dict] | None = None,
    on_event: EventCallback | None = None,
    chat_id: str | None = None,
    task_id: str | None = None,
    max_iterations: int = 15,
    extra_system: str = "",
    attachments: list[dict] | None = None,
    chat_artifacts: list[dict] | None = None,
) -> dict[str, Any]:
    emit = on_event or _noop

    # Classify intent and inject playbook guidance
    task_type = classify_intent(user_message)
    playbook_guidance = get_playbook_guidance(task_type) if task_type != "general" else ""
    if playbook_guidance:
        log.info(f"[agent] Playbook detected: {task_type} ({get_playbook_name(task_type)})")

    # Load cached query results from previous turns (same chat)
    query_results: dict[str, dict] = {}
    if chat_id:
        cached = _get_cached_query_results(chat_id)
        if cached:
            query_results = cached
            log.info(f"[agent] Loaded {len(cached)} cached queries for chat {chat_id}: {list(cached.keys())}")

    # Build system prompt with playbook + artifact context
    artifact_ctx = _build_artifact_context(chat_artifacts)
    extra_parts = [p for p in [extra_system, playbook_guidance, artifact_ctx] if p]
    full_extra = "\n\n".join(extra_parts)
    system_blocks = _build_system_prompt(full_extra, location_id)

    messages: list[dict] = [{"role": "system", "content": system_blocks}]
    if conversation_history:
        messages.extend(conversation_history)

    # Inject state summary before user message so LLM knows what's available
    state_summary = _build_state_summary(query_results, chat_artifacts)
    user_content = _build_user_content(user_message, attachments)
    if state_summary:
        if isinstance(user_content, str):
            user_content = f"{state_summary}\n\n---\nUSER MESSAGE: {user_content}"
        else:
            # Multimodal — prepend state as text block
            user_content.insert(0, {"type": "text", "text": state_summary + "\n\n---\nUSER MESSAGE:"})
    messages.append({"role": "user", "content": user_content})

    query_counter = 0
    sources_collected: list[dict] = []
    artifacts: list[dict] = []
    usage_records: list[dict] = []
    code_retry_counts: dict[str, int] = {}

    emit("agent_start", {"question": user_message, "model": model_id})

    for iteration in range(max_iterations):
        if chat_id and is_cancelled(chat_id):
            raise CancelledError(f"Chat {chat_id} cancelled")
        if task_id and is_cancelled(task_id):
            raise CancelledError(f"Task {task_id} cancelled")

        # Compress older tool results to save context tokens
        if iteration >= _COMPRESS_AFTER_ITERATION:
            messages = _compress_messages(messages)

        if iteration == 0:
            emit("thinking", {"step": 1, "message": "Analizando tu pregunta..."})
        else:
            emit("thinking", {"step": iteration + 1, "message": "Procesando resultados..."})

        response = traced_completion(
            model_id=model_id,
            messages=messages,
            step=f"agent_iter_{iteration + 1}",
            chat_id=chat_id,
            task_id=task_id,
            location_id=location_id,
            tools=UNIFIED_TOOLS,
            temperature=0.1,
        )

        u = getattr(response, "usage", None)
        prompt_tokens = getattr(u, "prompt_tokens", 0) or 0
        completion_tokens = getattr(u, "completion_tokens", 0) or 0
        total_tokens = getattr(u, "total_tokens", 0) or 0
        cache_read = getattr(u, "cache_read_input_tokens", 0) or 0
        cache_creation = getattr(u, "cache_creation_input_tokens", 0) or 0
        usage_records.append({
            "model": model_id,
            "step": f"agent_iter_{iteration + 1}",
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
            "total_tokens": total_tokens,
            "cache_read_tokens": cache_read,
            "cache_creation_tokens": cache_creation,
        })

        if cache_read > 0:
            log.info(f"[iter_{iteration+1}] Cache hit: {cache_read} tokens read from cache")

        if task_id:
            from hackathon_backend.services.lambdas.agent.core.chat_store import _estimate_cost
            from hackathon_backend.services.lambdas.agent.core.task_manager import add_task_cost, check_budget
            cost = _estimate_cost(model_id, prompt_tokens, completion_tokens, cache_read, cache_creation)
            add_task_cost(task_id, total_tokens, cost)
            budget = check_budget(task_id)
            if not budget.get("ok"):
                return {
                    "answer": f"Presupuesto agotado: {budget.get('reason', '')}",
                    "chart": None,
                    "sources": sources_collected,
                    "artifacts": artifacts,
                    "usage": usage_records,
                }

        choice = response.choices[0]

        # Agent finished — no more tool calls
        if choice.finish_reason == "stop" or not choice.message.tool_calls:
            final_text = choice.message.content or ""
            emit("agent_done", {"message": "Respuesta generada"})
            result = _parse_final_response(final_text, sources_collected)
            result["artifacts"] = artifacts
            result["usage"] = usage_records
            log.info(f"[run_agent] DONE iterations={iteration+1}, artifacts={[a.get('filename') for a in artifacts]}, answer_len={len(final_text)}")
            return result

        # Process tool calls
        messages.append(choice.message)

        tool_names = [tc.function.name for tc in choice.message.tool_calls]
        emit("tool_calls", {
            "message": f"Ejecutando: {', '.join(tool_names)}",
            "tools": tool_names,
            "iteration": iteration + 1,
        })

        for _tc in choice.message.tool_calls:
            _tc_args = _tc.function.arguments or "{}"
            log.info(f"[agent] Tool call: {_tc.function.name}({_tc_args[:500]})")
            emit("tool_call_detail", {
                "tool": _tc.function.name,
                "args": json.loads(_tc_args) if _tc_args else {},
                "iteration": iteration + 1,
            })

        for tool_call in choice.message.tool_calls:
            fn_name = tool_call.function.name
            tc_id = tool_call.id
            try:
                args = json.loads(tool_call.function.arguments)
            except json.JSONDecodeError:
                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps({"error": "Invalid JSON"}),
                })
                continue

            # ---------------------------------------------------------------
            # dynamo_query — returns dataset card
            # ---------------------------------------------------------------
            if fn_name == "dynamo_query":
                query_counter += 1
                query_key = f"query_{query_counter}"
                table_name = args["table_name"]
                table_label = table_name.replace("_", " ")

                query_details = {
                    "table": table_name, "query_key": query_key,
                    "index": args.get("index_name"),
                    "pk_field": args.get("pk_field", "userId"),
                    "sk_field": args.get("sk_field"),
                    "sk_condition": args.get("sk_condition"),
                    "filter": args.get("filter_expression"),
                    "limit": args.get("limit"),
                }
                emit("querying", {"message": f"Consultando {table_label}...", **query_details})
                log.info(f"[dynamo_query] {query_key}: {json.dumps(query_details, default=str)}")

                result = _execute_query(
                    table_name=table_name,
                    location_id=location_id,
                    index_name=args.get("index_name"),
                    pk_field=args.get("pk_field", "userId"),
                    pk_value=args.get("pk_value"),
                    sk_field=args.get("sk_field"),
                    sk_condition=args.get("sk_condition"),
                    filter_expression=args.get("filter_expression"),
                    limit=args.get("limit"),
                )
                query_results[query_key] = result
                _cache_query_results(chat_id, query_results)

                if result.get("success"):
                    emit("query_result", {
                        "query_key": query_key, "table": table_name,
                        "count": result["count"],
                        "message": f"Encontrados {result['count']} registros en {table_label}",
                        **query_details,
                    })
                    log.info(f"[dynamo_query] {query_key}: {result['count']} items from {table_name}")

                    # Collect sources for invoice/bank tables
                    if table_name in ("User_Expenses", "User_Invoice_Incomes"):
                        for item in result["items"]:
                            src = _extract_source(item)
                            if src:
                                sources_collected.append(src)
                    elif table_name == "Bank_Reconciliations":
                        for item in result["items"]:
                            sources_collected.append({
                                "categoryDate": item.get("SK", ""),
                                "supplier": item.get("merchant") or item.get("description", ""),
                                "invoice_date": item.get("bookingDate"),
                                "total": item.get("amount"),
                                "reconciled": bool(item.get("reconciled")) or item.get("status") == "MATCHED",
                                "category": "BANK",
                                "matched_expense_id": item.get("matched_expense_id"),
                                "matched_invoice_id": item.get("matched_invoice_id"),
                                "match_type": item.get("match_type"),
                                "transactionId": item.get("transactionId"),
                            })

                # Build dataset card instead of raw items
                if result.get("success"):
                    card = _build_dataset_card(query_key, result, table_name)
                    response_for_llm = card
                else:
                    response_for_llm = {
                        "query_key": query_key, "success": False,
                        "table": table_name,
                        "error": result.get("error", "Unknown"),
                    }

                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps(response_for_llm, ensure_ascii=False, default=str),
                })

            # ---------------------------------------------------------------
            # run_code — sandboxed local execution (unified: analysis + files)
            # ---------------------------------------------------------------
            elif fn_name == "run_code":
                code = args.get("code", "")
                file_task_id = task_id or f"chat_{str(uuid.uuid4())[:8]}"

                # If no data and code references data[], tell the LLM to query first
                # But allow run_code without data for file editing (uses existing_files)
                if not query_results and "data[" in code:
                    log.warning(f"[run_code] No data available — telling LLM to query first")
                    messages.append({
                        "role": "tool", "tool_call_id": tc_id,
                        "content": json.dumps({
                            "error": "NO DATA AVAILABLE. The data dict is empty — you must call dynamo_query first to fetch data before calling run_code. Call dynamo_query now.",
                            "hint": "Call dynamo_query to fetch the data you need, then call run_code.",
                        }),
                    })
                    continue

                code_preview = code.strip().split("\n")[0][:80]
                log.info(f"[run_code] Code:\n{code}")
                emit("analyzing", {
                    "message": "Ejecutando codigo...",
                    "detail": code_preview,
                    "code": code[:200],
                    "available_queries": list(query_results.keys()),
                    "query_counts": {k: v.get("count", 0) for k, v in query_results.items()},
                })

                # Combine artifacts from this turn + previous turns
                all_artifacts = (chat_artifacts or []) + artifacts
                exec_result = _safe_exec(code, query_results, file_task_id,
                                        existing_artifacts=all_artifacts)
                elapsed_ms = exec_result.get("elapsed_ms", 0)

                # Audit every execution
                _audit_code_execution(
                    task_id=file_task_id, code=code, query_results=query_results,
                    result=exec_result.get("result"), error=exec_result.get("error"),
                    elapsed_ms=elapsed_ms, files=exec_result.get("files"),
                    location_id=location_id, chat_id=chat_id,
                )

                if not exec_result["success"]:
                    # Retry tracking (global counter, not per-iteration)
                    retry_key = "run_code_failures"
                    code_retry_counts[retry_key] = code_retry_counts.get(retry_key, 0) + 1

                    if code_retry_counts[retry_key] >= 3:
                        log.error(f"[run_code] 3 failures, giving up: {exec_result['error']}")
                        messages.append({
                            "role": "tool", "tool_call_id": tc_id,
                            "content": json.dumps({
                                "error": exec_result["error"],
                                "fatal": True,
                                "message": "Code execution failed 3 times. Explain the issue to the user.",
                            }),
                        })
                    else:
                        messages.append({
                            "role": "tool", "tool_call_id": tc_id,
                            "content": json.dumps({
                                "error": exec_result["error"],
                                "attempt": code_retry_counts[retry_key],
                                "max_attempts": 3,
                                "hint": "Fix the code and call run_code again.",
                            }),
                        })
                    continue

                # Success — check for result and/or files
                result_val = exec_result.get("result")
                generated_files = exec_result.get("files", [])

                # Validate generated files — catch empty outputs
                if generated_files:
                    generated_files = _validate_generated_files(generated_files)
                    # If ALL files are invalid (empty), treat as failure
                    if generated_files and all(not f.get("valid", True) for f in generated_files):
                        retry_key = "run_code_failures"
                        code_retry_counts[retry_key] = code_retry_counts.get(retry_key, 0) + 1
                        warnings = "; ".join(f.get("warning", "empty file") for f in generated_files)
                        log.warning(f"[run_code] All generated files are empty: {warnings}")
                        if code_retry_counts[retry_key] >= 3:
                            messages.append({
                                "role": "tool", "tool_call_id": tc_id,
                                "content": json.dumps({
                                    "error": f"Files generated but ALL are empty: {warnings}",
                                    "fatal": True,
                                    "message": "Generated files have no data after 3 attempts. Check your data source — you may be reading from the wrong query key. Tell the user what went wrong.",
                                }),
                            })
                            continue
                        else:
                            # Tell LLM which queries have data so it can fix
                            data_summary = {k: v.get("count", len(v.get("items", [])))
                                            for k, v in query_results.items()}
                            messages.append({
                                "role": "tool", "tool_call_id": tc_id,
                                "content": json.dumps({
                                    "error": f"Files generated but ALL are empty: {warnings}",
                                    "attempt": code_retry_counts[retry_key],
                                    "max_attempts": 3,
                                    "available_data": data_summary,
                                    "hint": "Your code produced empty files. Check which query key has data — look at available_data to find the correct key with items > 0. Fix the data source and try again.",
                                }),
                            })
                            continue

                # Handle files → artifacts
                generated_filenames = []
                from hackathon_backend.services.lambdas.agent.core.storage import (
                    _use_s3 as _s3_check, save_artifact as _save_art,
                )
                log.info(f"[run_code] {len(generated_files)} files to process, s3={_s3_check()}, task={file_task_id}")
                for f in generated_files:
                    file_url = f"/api/tasks/{file_task_id}/artifacts/{f['filename']}"
                    if _s3_check():
                        fp = f["path"]
                        if os.path.isfile(fp):
                            log.info(f"[run_code] Uploading to S3: {f['filename']} ({f.get('size_bytes', 0)} bytes)")
                            try:
                                with open(fp, "rb") as _fh:
                                    s3_res = _save_art(file_task_id, f["filename"], _fh.read())
                                file_url = s3_res.get("url", file_url)
                                log.info(f"[run_code] S3 upload OK: {f['filename']} → url_len={len(file_url)}")
                            except Exception as s3_err:
                                log.error(f"[run_code] S3 upload FAILED for {f['filename']}: {s3_err}")
                        else:
                            log.warning(f"[run_code] File not found on disk: {fp}")
                    artifact = {
                        "filename": f["filename"], "path": f["path"],
                        "task_id": file_task_id,
                        "type": f.get("type", "file"),
                        "size_bytes": f.get("size_bytes", 0),
                        "url": file_url,
                    }
                    artifacts.append(artifact)
                    generated_filenames.append(f["filename"])

                    if task_id:
                        from hackathon_backend.services.lambdas.agent.core.task_manager import add_task_artifact
                        add_task_artifact(task_id, artifact)

                if generated_filenames:
                    emit("file_generated", {
                        "message": f"Archivo generado: {', '.join(generated_filenames)}",
                        "files": generated_filenames,
                        "success": True,
                    })

                # If result has 'answer' AND we have generated files, return immediately
                # (this means the agent completed the final step with Excel/report).
                # If no files, let the LLM continue — it may have more steps to do.
                if isinstance(result_val, dict) and "answer" in result_val and generated_files:
                    log.info(f"[run_code] Final answer with files: {str(result_val.get('answer', ''))[:200]}")
                    emit("analysis_result", {
                        "message": "Analisis completado",
                        "answer_preview": str(result_val.get("answer", ""))[:300],
                        "has_chart": result_val.get("chart") is not None,
                        "sources_count": len(result_val.get("sources") or sources_collected),
                    })
                    emit("agent_done", {"message": "Analisis completado"})
                    ret = {
                        "answer": result_val.get("answer", ""),
                        "chart": result_val.get("chart"),
                        "sources": result_val.get("sources") or sources_collected,
                        "artifacts": artifacts,
                        "usage": usage_records,
                    }
                    # Pass through extra fields (todo, close_status, etc.)
                    for extra_key in ("todo", "close_status", "kpis"):
                        if result_val.get(extra_key):
                            ret[extra_key] = result_val[extra_key]
                    return ret

                # Build tool response
                tool_response: dict[str, Any] = {"success": True}
                if result_val is not None:
                    tool_response["result"] = result_val
                if generated_files:
                    tool_response["files"] = [
                        {
                            "filename": f["filename"],
                            "url": f"/api/tasks/{file_task_id}/artifacts/{f['filename']}",
                            "valid": f.get("valid", True),
                            "metrics": f.get("metrics", {}),
                            **({"warning": f["warning"]} if f.get("warning") else {}),
                        }
                        for f in generated_files
                    ]
                tool_response["elapsed_ms"] = elapsed_ms

                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps(tool_response, ensure_ascii=False, default=str),
                })

            # ---------------------------------------------------------------
            # dispatch_subagents — parallel document processing
            # ---------------------------------------------------------------
            elif fn_name == "dispatch_subagents":
                subtasks = args.get("subtasks", [])
                max_parallel = min(args.get("max_parallel", 5), 10)

                if not subtasks:
                    messages.append({
                        "role": "tool", "tool_call_id": tc_id,
                        "content": json.dumps({"error": "No subtasks provided."}),
                    })
                    continue

                total_docs = sum(len(st.get("documents", [])) for st in subtasks)
                log.info(f"[dispatch_subagents] {len(subtasks)} subtasks, {total_docs} docs, max_parallel={max_parallel}")
                emit("dispatching_subagents", {
                    "message": f"Lanzando {len(subtasks)} subagentes para procesar {total_docs} documentos...",
                    "subtask_count": len(subtasks),
                    "total_documents": total_docs,
                })

                dispatch_result = _dispatch_subagents(
                    subtasks=subtasks,
                    location_id=location_id,
                    model_id=model_id,
                    chat_id=chat_id,
                    task_id=task_id,
                    max_parallel=max_parallel,
                    on_event=emit,
                )

                # Collect artifacts and usage from subagents
                for sr in dispatch_result.get("subagent_results", []):
                    for art in sr.get("artifacts", []):
                        artifacts.append(art)
                    usage_records.extend(sr.get("usage", []))

                # Store FULL subagent results in query_results so run_code can access them
                # (keeps them OUT of the LLM context but available for code execution)
                all_extracted = []
                subagent_summaries = []
                for sr in dispatch_result.get("subagent_results", []):
                    sa_id = sr["subagent_id"]
                    sa_summary = {
                        "subagent_id": sa_id,
                        "success": sr["success"],
                        "cost_usd": sr["cost_usd"],
                        "iterations": sr["iterations"],
                    }
                    if sr.get("error"):
                        sa_summary["error"] = sr["error"]

                    # Extract compact summary for LLM (not full data)
                    result_data = sr.get("result") or {}
                    if isinstance(result_data, dict):
                        extracted = result_data.get("extracted_data", [])
                        all_extracted.extend(extracted)
                        sa_summary["documents_processed"] = len(extracted)
                        sa_summary["total_amount"] = sum(
                            d.get("total", 0) or 0 for d in extracted if isinstance(d, dict)
                        )
                        # Compact preview: filename + total + matched status
                        sa_summary["documents"] = [
                            {
                                "filename": d.get("filename", "?"),
                                "supplier": d.get("supplier", "?"),
                                "total": d.get("total"),
                                "date": d.get("date"),
                                "matched": d.get("matched_db_record") is not None,
                                "issues": d.get("issues", []),
                            }
                            for d in extracted[:20]  # cap at 20 per subagent
                        ]
                        if result_data.get("issues"):
                            sa_summary["issues"] = result_data["issues"]
                    subagent_summaries.append(sa_summary)

                # Store full extracted data in query_results for run_code access
                query_counter += 1
                subagent_data_key = f"subagent_results"
                query_results[subagent_data_key] = {
                    "items": all_extracted,
                    "count": len(all_extracted),
                    "table": "subagent_extracted_data",
                    "success": True,
                }
                _cache_query_results(chat_id, query_results)

                # Emit detailed progress for frontend
                for sa_sum in subagent_summaries:
                    emit("subagent_result", {
                        "subagent_id": sa_sum["subagent_id"],
                        "success": sa_sum["success"],
                        "documents_processed": sa_sum.get("documents_processed", 0),
                        "total_amount": sa_sum.get("total_amount", 0),
                        "cost_usd": sa_sum.get("cost_usd", 0),
                        "documents": sa_sum.get("documents", []),
                    })

                # Build COMPACT tool response for LLM (not the full data)
                tool_response = {
                    "success": dispatch_result["success"],
                    "summary": dispatch_result["summary"],
                    "total_cost_usd": dispatch_result["total_cost_usd"],
                    "documents_processed": dispatch_result["documents_processed"],
                    "total_extracted": len(all_extracted),
                    "subagent_results": subagent_summaries,
                    "data_access": (
                        f"Full extracted data stored in data['{subagent_data_key}']['items'] "
                        f"({len(all_extracted)} documents). Use run_code to process them."
                    ),
                }

                emit("subagents_done", {
                    "message": dispatch_result["summary"],
                    "success": dispatch_result["success"],
                    "subagent_count": len(subtasks),
                    "documents_processed": len(all_extracted),
                    "total_cost": dispatch_result["total_cost_usd"],
                    "elapsed_s": dispatch_result.get("elapsed_s"),
                    "subagents": subagent_summaries,
                })

                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps(tool_response, ensure_ascii=False, default=str),
                })

            # ---------------------------------------------------------------
            # edit_file — DEPRECATED, redirect to run_code
            # ---------------------------------------------------------------
            elif fn_name == "edit_file":
                log.warning(f"[agent] LLM called deprecated edit_file tool, redirecting to run_code")
                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps({
                        "error": "edit_file tool has been removed. Use run_code instead. "
                                 "The existing_files dict contains previously generated files. "
                                 "Load the file with openpyxl.load_workbook(existing_files['filename.xlsx']), "
                                 "apply your edits, and save to f'{output_dir}/filename.xlsx'.",
                    }),
                })

            else:
                messages.append({
                    "role": "tool", "tool_call_id": tc_id,
                    "content": json.dumps({"error": f"Unknown tool: {fn_name}"}),
                })

    # Max iterations
    return {
        "answer": "Se ha alcanzado el limite de iteraciones.",
        "chart": None,
        "sources": sources_collected,
        "artifacts": artifacts,
        "usage": usage_records,
    }


# ---------------------------------------------------------------------------
# Response parser
# ---------------------------------------------------------------------------
def _parse_final_response(text: str, default_sources: list[dict]) -> dict:
    result = {"answer": text, "chart": None, "sources": default_sources}
    if "```json" in text:
        try:
            start = text.index("```json") + 7
            end = text.index("```", start)
            parsed = json.loads(text[start:end].strip())
            if "answer" in parsed:
                result["answer"] = parsed["answer"]
            if parsed.get("chart"):
                result["chart"] = parsed["chart"]
            if parsed.get("sources"):
                result["sources"] = parsed["sources"]
            clean = (text[:start - 7] + text[end + 3:]).strip()
            if clean and not result.get("answer"):
                result["answer"] = clean
        except (ValueError, json.JSONDecodeError):
            pass
    return result


# ---------------------------------------------------------------------------
# Detect if a message needs background processing (heavy task)
# ---------------------------------------------------------------------------
# Maps playbook task_type → task_executor task_type (for heavy background tasks)
_HEAVY_PLAYBOOK_TYPES: dict[str, str] = {
    "prediccion_cashflow": "cash_flow_forecast",
    "reportes_financieros": "pack_reporting",
    "auditoria_iva": "modelo_303",
    "conciliacion": "bank_reconciliation",
}


def detect_heavy_task(user_message: str) -> str | None:
    """
    Detect if a message requires heavy background processing.

    Uses semantic intent classification (via cheap LLM call) instead of keywords.
    Returns the task_type string if heavy, or None for inline processing.
    """
    task_type = classify_intent(user_message)

    # Explanations and general questions are never heavy
    if task_type in ("explicacion_humana", "general"):
        return None

    # Only certain task types are heavy enough for background processing
    return _HEAVY_PLAYBOOK_TYPES.get(task_type)
