"""
Integration test suite for TaskAgent — runs all 5 hackathon use cases
against real DynamoDB data, grades outputs, saves artifacts for review.

Usage:
    python -m scripts.test_task_agent
    python -m scripts.test_task_agent --query 1
    python -m scripts.test_task_agent --runs 2
    python -m scripts.test_task_agent --user-id deloitte-84
    python -m scripts.test_task_agent --layer1          # unit tests only (no LLM)
    python -m scripts.test_task_agent --consistency      # run each 3x, compare

Output:
    test_output/task_agent/q{N}_{name}/
        result.json, trace.json, answer.md, chart.html, grade.json
    test_output/task_agent/summary.json
    test_output/task_agent/report.html               (visual report)
"""
from __future__ import annotations

import argparse
import json
import logging
import math
import os
import sys
import time
import traceback

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
)
logger = logging.getLogger("task_agent_test")

# ---------------------------------------------------------------------------
# Test queries — the 5 hackathon use cases
# ---------------------------------------------------------------------------

TASK_QUERIES = [
    {
        "id": 1,
        "name": "cash_flow_forecast",
        "query": "Genera una previsión de tesorería (cash flow forecast) para las próximas 13 semanas",
        "expected": {
            "has_answer": True,
            "has_data": True,
            "min_sources": 3,
            "has_chart": True,
            "has_export": True,
            "answer_language": "es",
        },
        "quality_keywords": [
            "semana",
            "saldo",
            "flujo",
        ],
    },
    {
        "id": 2,
        "name": "monthly_pnl_report",
        "query": "Prepara el pack de reporting mensual con P&L y KPIs principales de diciembre 2025",
        "expected": {
            "has_answer": True,
            "has_data": True,
            "min_sources": 2,
            "has_chart": True,
            "has_export": True,
            "answer_language": "es",
        },
        "quality_keywords": [
            "ingreso",
            "gasto",
            "margen",
        ],
    },
    {
        "id": 3,
        "name": "modelo_303_vat",
        "query": "Genera el borrador del Modelo 303 (IVA) del Q4 2025",
        "expected": {
            "has_answer": True,
            "has_data": True,
            "min_sources": 1,
            "has_export": True,
            "answer_language": "es",
        },
        "quality_keywords": [
            "IVA",
            "repercutido",
            "soportado",
        ],
    },
    {
        "id": 4,
        "name": "aging_analysis",
        "query": "Análisis de antigüedad (aging) de cobros y pagos pendientes",
        "expected": {
            "has_answer": True,
            "has_data": True,
            "min_sources": 1,
            "has_chart": True,
            "has_export": True,
            "answer_language": "es",
        },
        "quality_keywords": [
            "30",
            "60",
            "90",
        ],
    },
    {
        "id": 5,
        "name": "client_profitability",
        "query": "Análisis de rentabilidad por cliente del último trimestre",
        "expected": {
            "has_answer": True,
            "has_data": True,
            "min_sources": 1,
            "has_chart": True,
            "has_export": True,
            "answer_language": "es",
        },
        "quality_keywords": [
            "cliente",
            "margen",
            "ingreso",
        ],
    },
]

# Spanish indicator words (for language check)
SPANISH_INDICATORS = [
    "el", "la", "los", "las", "de", "del", "en", "por", "para",
    "total", "resultado", "datos", "periodo",
]


# ---------------------------------------------------------------------------
# Layer 1: Tool handler unit tests (no LLM needed)
# ---------------------------------------------------------------------------

def run_layer1_tests() -> list[dict]:
    """Test individual tool handlers without running the full agent or LLM."""
    from hackathon_backend.agents.task_agent import TaskAgent
    from hackathon_backend.agents.export_tool import generate_export

    base_output = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "test_output", "task_agent", "layer1",
    )
    os.makedirs(base_output, exist_ok=True)

    results = []

    # Test 1: get_current_date
    print("\n  [Layer1] get_current_date")
    agent = TaskAgent(user_id="test-user")
    d = agent._handle_get_current_date()
    checks = {
        "has_today": "today" in d,
        "has_quarter": "current_quarter" in d and d["current_quarter"].startswith("Q"),
        "has_boundaries": all(k in d for k in ("quarter_start", "quarter_end", "fiscal_year_start")),
        "date_format": len(d.get("today", "")) == 10,
    }
    ok = all(checks.values())
    results.append({"test": "get_current_date", "passed": ok, "checks": checks})
    print(f"    {'PASS' if ok else 'FAIL'}: {checks}")

    # Test 2: CSV export
    print("\n  [Layer1] generate_export CSV")
    rows = [{"month": "2026-01", "total": 1234.56}, {"month": "2026-02", "total": 7890.12}]
    csv_r = generate_export(json.dumps(rows), fmt="csv", filename="layer1_csv", output_dir=base_output)
    csv_ok = csv_r.get("success") and os.path.exists(csv_r.get("file_path", ""))
    if csv_ok:
        import csv as _csv
        with open(csv_r["file_path"], "r", encoding="utf-8-sig") as f:
            parsed = list(_csv.DictReader(f))
            csv_ok = len(parsed) == 2 and "month" in parsed[0]
    results.append({"test": "csv_export", "passed": csv_ok, "checks": csv_r})
    print(f"    {'PASS' if csv_ok else 'FAIL'}: rows={csv_r.get('row_count')}, path={csv_r.get('file_path','')}")

    # Test 3: XLSX export (multi-sheet, styled)
    print("\n  [Layer1] generate_export XLSX")
    sheets = {"Sheet1": rows, "Summary": [{"metric": "Total", "value": 9124.68}]}
    xlsx_r = generate_export(json.dumps(sheets), fmt="xlsx", filename="layer1_xlsx", output_dir=base_output, title="Test")
    xlsx_ok = xlsx_r.get("success") and os.path.exists(xlsx_r.get("file_path", ""))
    if xlsx_ok:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_r["file_path"])
            xlsx_ok = len(wb.sheetnames) == 2
            # Check header styling
            ws = wb[wb.sheetnames[0]]
            header_cell = ws.cell(row=3, column=1)  # row 3 because title is row 1
            xlsx_ok = xlsx_ok and header_cell.font.bold
            wb.close()
        except Exception as exc:
            xlsx_ok = False
    results.append({"test": "xlsx_export", "passed": xlsx_ok, "checks": xlsx_r})
    print(f"    {'PASS' if xlsx_ok else 'FAIL'}: sheets={xlsx_r.get('sheets')}, rows={xlsx_r.get('row_count')}")

    # Test 4: ask_user without callback (graceful fallback)
    print("\n  [Layer1] ask_user (no callback)")
    ask_r = agent._handle_ask_user("Test?", options=["A", "B"])
    ask_ok = "user_response" in ask_r and "judgment" in ask_r["user_response"].lower()
    results.append({"test": "ask_user_fallback", "passed": ask_ok, "checks": ask_r})
    print(f"    {'PASS' if ask_ok else 'FAIL'}: {ask_r}")

    return results


# ---------------------------------------------------------------------------
# Export validation helper
# ---------------------------------------------------------------------------

def _validate_export_files(exports: list[str]) -> bool:
    """Check that exported files are valid and parseable."""
    for path in exports:
        if not os.path.exists(path):
            return False
        if path.endswith(".csv"):
            try:
                import csv as _csv
                with open(path, "r", encoding="utf-8-sig") as f:
                    rows = list(_csv.reader(f))
                if len(rows) < 2:  # header + at least 1 data row
                    return False
            except Exception:
                return False
        elif path.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path)
                if not wb.sheetnames:
                    return False
                wb.close()
            except Exception:
                return False
    return True


def grade_result(result, query_info: dict, output_dir: str) -> dict:
    """Grade an AgentResult against expected criteria. Returns grade dict."""
    expected = query_info["expected"]
    quality_kw = query_info.get("quality_keywords", [])
    grade = {"total": 0, "max": 100, "checks": {}}

    def check(name: str, passed: bool, points: int, detail: str = ""):
        grade["checks"][name] = {
            "passed": passed,
            "points": points if passed else 0,
            "max_points": points,
            "detail": detail,
        }
        if passed:
            grade["total"] += points

    # 1. Success (10 pts)
    check("success", result.success, 10,
          f"success={result.success}, error={result.error}")

    if not result.success:
        # Can't grade further if failed
        grade["max"] = 10
        return grade

    data = result.data or {}
    answer = str(data.get("answer", ""))
    sources = data.get("sources", [])
    metrics = data.get("metrics", {})
    exports = data.get("exports", [])

    # 2. Answer non-empty + language (10 pts)
    has_answer = len(answer) > 50
    answer_lower = answer.lower()
    spanish_count = sum(1 for w in SPANISH_INDICATORS if f" {w} " in f" {answer_lower} ")
    is_spanish = spanish_count >= 3
    check("answer_quality", has_answer and is_spanish, 10,
          f"len={len(answer)}, spanish_words={spanish_count}")

    # 3. Structured data present (10 pts)
    has_data = data.get("data") is not None
    check("structured_data", has_data, 10,
          f"data={'present' if has_data else 'missing'}")

    # 4. Metrics exist (15 pts)
    has_metrics = len(metrics) > 0
    check("metrics_exist", has_metrics, 15,
          f"metrics_keys={list(metrics.keys())}")

    # 5. Metrics sanity (10 pts)
    metrics_sane = True
    bad_metrics = []
    for k, v in metrics.items():
        if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
            metrics_sane = False
            bad_metrics.append(f"{k}={v}")
    check("metrics_sanity", metrics_sane, 10,
          f"bad_metrics={bad_metrics}" if bad_metrics else "all sane")

    # 6. Sources count (10 pts)
    min_sources = expected.get("min_sources", 1)
    check("sources_count", len(sources) >= min_sources, 10,
          f"sources={len(sources)}, min={min_sources}")

    # 7. Chart (10 pts)
    if expected.get("has_chart"):
        has_chart = result.chart_html is not None and len(result.chart_html or "") > 100
        check("chart_generated", has_chart, 10,
              f"chart_html_len={len(result.chart_html or '')}")
    else:
        check("chart_generated", True, 10, "chart not expected")

    # 8. Export generated (5 pts) + valid (5 pts)
    if expected.get("has_export"):
        has_export = len(exports) > 0 and any(os.path.exists(p) for p in exports)
        check("export_generated", has_export, 5,
              f"exports={exports}")
        if has_export:
            exports_valid = _validate_export_files(exports)
            check("export_valid", exports_valid, 5,
                  f"all files valid" if exports_valid else "some files invalid")
        else:
            check("export_valid", False, 5, "no exports to validate")
    else:
        check("export_generated", True, 5, "export not expected")
        check("export_valid", True, 5, "export not expected")

    # 9. Tool calls used (5 pts)
    check("tool_calls_used", result.iterations_used >= 3, 5,
          f"tool_calls={result.iterations_used}")

    # 10. Quality keywords (10 pts)
    kw_found = [kw for kw in quality_kw if kw.lower() in answer_lower]
    kw_ratio = len(kw_found) / max(len(quality_kw), 1)
    check("quality_keywords", kw_ratio >= 0.5, 10,
          f"found={kw_found}, expected={quality_kw}")

    grade["passed"] = grade["total"] >= 70
    return grade


def save_artifacts(result, query_info: dict, grade: dict, output_dir: str, elapsed: float):
    """Save all test artifacts for manual review."""
    os.makedirs(output_dir, exist_ok=True)

    data = result.data or {}

    # result.json
    with open(os.path.join(output_dir, "result.json"), "w") as f:
        json.dump({
            "success": result.success,
            "data": data,
            "error": result.error,
            "iterations_used": result.iterations_used,
            "elapsed_seconds": elapsed,
        }, f, indent=2, default=str, ensure_ascii=False)

    # trace.json
    with open(os.path.join(output_dir, "trace.json"), "w") as f:
        json.dump(result.trace, f, indent=2, default=str, ensure_ascii=False)

    # answer.md
    answer = data.get("answer", result.error or "No answer")
    with open(os.path.join(output_dir, "answer.md"), "w") as f:
        f.write(f"# {query_info['name']}\n\n")
        f.write(f"**Query**: {query_info['query']}\n\n")
        f.write(f"**Score**: {grade['total']}/100 {'PASS' if grade.get('passed') else 'FAIL'}\n\n")
        f.write(f"**Time**: {elapsed:.1f}s | **Tool calls**: {result.iterations_used}\n\n")
        f.write("---\n\n")
        f.write(str(answer))

    # chart.html
    if result.chart_html:
        full_html = (
            f'<!DOCTYPE html><html><head><meta charset="utf-8">'
            f'<title>{query_info["name"]}</title></head>'
            f'<body style="background:#f5f5f5;font-family:sans-serif;padding:20px">'
            f'<h2>{query_info["name"]}</h2>'
            f'{result.chart_html}'
            f'</body></html>'
        )
        with open(os.path.join(output_dir, "chart.html"), "w") as f:
            f.write(full_html)

    # grade.json
    with open(os.path.join(output_dir, "grade.json"), "w") as f:
        json.dump(grade, f, indent=2, ensure_ascii=False)


def run_single_test(query_info: dict, user_id: str, base_output_dir: str,
                    model_id: str = "claude-sonnet-4.5") -> dict:
    """Run a single TaskAgent test query. Returns summary dict."""
    from hackathon_backend.agents.task_agent import TaskAgent

    qid = query_info["id"]
    name = query_info["name"]
    output_dir = os.path.join(base_output_dir, f"q{qid}_{name}")

    print(f"\n{'='*70}")
    print(f"  Q{qid} [{name}]: {query_info['query'][:60]}...")
    print(f"  Model: {model_id}")
    print(f"{'='*70}")

    def progress_cb(event: str, data: dict):
        desc = data.get("description", data.get("plan", data.get("step", "")))
        print(f"    [{event}] {str(desc)[:80]}")

    agent = TaskAgent(
        user_id=user_id,
        model_id=model_id,
        progress_callback=progress_cb,
        export_dir=os.path.join(base_output_dir, "exports"),
    )

    start = time.time()
    try:
        result = agent.run(query_info["query"])
    except Exception as exc:
        logger.error("Q%d EXCEPTION: %s", qid, exc)
        traceback.print_exc()
        from hackathon_backend.agents.agent import AgentResult
        result = AgentResult(success=False, error=str(exc))
    elapsed = time.time() - start

    # Grade
    grade = grade_result(result, query_info, output_dir)

    # Save artifacts
    save_artifacts(result, query_info, grade, output_dir, elapsed)

    # Print summary
    status = "PASS" if grade.get("passed") else "FAIL"
    print(f"\n  Score: {grade['total']}/100 [{status}] ({elapsed:.1f}s)")
    for check_name, check_info in grade["checks"].items():
        marker = "+" if check_info["passed"] else "X"
        print(f"    [{marker}] {check_name}: {check_info['points']}/{check_info['max_points']} — {check_info['detail'][:60]}")

    return {
        "id": qid,
        "name": name,
        "score": grade["total"],
        "passed": grade.get("passed", False),
        "elapsed": elapsed,
        "tool_calls": result.iterations_used,
        "error": result.error,
    }


def generate_html_report(all_results: list[dict], base_output: str):
    """Generate an HTML report with scores, charts inline, and export links."""
    parts = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        "<title>TaskAgent Test Report</title>",
        "<style>",
        "body{font-family:system-ui;max-width:1200px;margin:auto;padding:20px;background:#f5f5f5}",
        "h1{color:#f2764b}h2{color:#333;border-bottom:2px solid #f2764b;padding-bottom:8px}",
        ".card{background:white;border-radius:8px;padding:20px;margin:16px 0;box-shadow:0 2px 4px rgba(0,0,0,0.1)}",
        ".pass{color:#22c55e;font-weight:bold}.fail{color:#ef4444;font-weight:bold}",
        "table{border-collapse:collapse;width:100%}th,td{border:1px solid #ddd;padding:8px;text-align:left}",
        "th{background:#f2764b;color:white}",
        "</style></head><body>",
        "<h1>TaskAgent Test Report</h1>",
    ]

    avg = sum(r["score"] for r in all_results) / max(len(all_results), 1)
    passes = sum(1 for r in all_results if r["passed"])
    parts.append(f"<div class='card'><h2>Summary: {avg:.0f}/100 avg, {passes}/{len(all_results)} passed</h2>")
    parts.append("<table><tr><th>Case</th><th>Score</th><th>Time</th><th>Tools</th><th>Status</th></tr>")
    for r in all_results:
        cls = "pass" if r["passed"] else "fail"
        parts.append(f"<tr><td>{r['name']}</td><td class='{cls}'>{r['score']}/100</td>"
                      f"<td>{r['elapsed']:.1f}s</td><td>{r['tool_calls']}</td>"
                      f"<td class='{cls}'>{'PASS' if r['passed'] else 'FAIL'}</td></tr>")
    parts.append("</table></div>")

    # Per-case: embed chart + link exports
    for r in all_results:
        case_dir = os.path.join(base_output, f"q{r['id']}_{r['name']}")
        parts.append(f"<div class='card'><h2>Q{r['id']}: {r['name']} — {r['score']}/100</h2>")

        # Answer preview
        answer_path = os.path.join(case_dir, "answer.md")
        if os.path.exists(answer_path):
            with open(answer_path) as f:
                answer_text = f.read()[:1000]
            parts.append(f"<pre style='white-space:pre-wrap;font-size:0.85em'>{answer_text}</pre>")

        # Inline chart
        chart_path = os.path.join(case_dir, "chart.html")
        if os.path.exists(chart_path):
            with open(chart_path) as f:
                chart_html = f.read()
            # Extract chart div+scripts
            div_start = chart_html.find("<div")
            script_end = chart_html.rfind("</script>")
            if div_start >= 0 and script_end > div_start:
                parts.append(chart_html[div_start:script_end + len("</script>")])

        # Export links
        export_dir = os.path.join(base_output, "exports")
        if os.path.isdir(export_dir):
            export_files = [f for f in os.listdir(export_dir) if r["name"] in f or f"q{r['id']}" in f]
            if export_files:
                parts.append("<p><strong>Exports:</strong> " + ", ".join(export_files) + "</p>")

        parts.append("</div>")

    parts.append("</body></html>")
    report_path = os.path.join(base_output, "report.html")
    with open(report_path, "w") as f:
        f.write("\n".join(parts))
    print(f"  HTML report: {report_path}")


def main():
    parser = argparse.ArgumentParser(description="TaskAgent integration test suite")
    parser.add_argument("--query", type=int, help="Run only query N (1-5)")
    parser.add_argument("--runs", type=int, default=1, help="Runs per query (default: 1)")
    parser.add_argument("--user-id", default="deloitte-84", help="User ID")
    parser.add_argument("--model", default="claude-sonnet-4.5",
                        help="Planning brain model (default: claude-sonnet-4.5 for speed)")
    parser.add_argument("--layer1", action="store_true", help="Run Layer 1 unit tests only (no LLM)")
    parser.add_argument("--consistency", action="store_true", help="Run each case multiple times and compare")
    args = parser.parse_args()

    print("Initializing models...")
    from hackathon_backend.services.lambdas.agent.core.config import init_all
    init_all()
    print("Models loaded.\n")

    # Layer 1 only mode
    if args.layer1:
        print("=" * 60)
        print("  Layer 1: Tool Handler Unit Tests (no LLM)")
        print("=" * 60)
        l1 = run_layer1_tests()
        passed = sum(1 for r in l1 if r["passed"])
        print(f"\n  Layer 1: {passed}/{len(l1)} passed")
        sys.exit(0 if passed == len(l1) else 1)

    queries = TASK_QUERIES
    if args.query:
        queries = [q for q in TASK_QUERIES if q["id"] == args.query]
        if not queries:
            print(f"Query {args.query} not found (valid: 1-5)")
            sys.exit(1)

    base_output = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "test_output",
        "task_agent",
    )
    os.makedirs(base_output, exist_ok=True)

    all_results = []

    for query_info in queries:
        for run_num in range(1, args.runs + 1):
            if args.runs > 1:
                print(f"\n  --- Run {run_num}/{args.runs} ---")
            summary = run_single_test(query_info, args.user_id, base_output, model_id=args.model)
            all_results.append(summary)

    # Overall summary
    print(f"\n{'='*70}")
    print(f"  OVERALL SUMMARY")
    print(f"{'='*70}")

    total_pass = sum(1 for r in all_results if r["passed"])
    total_fail = sum(1 for r in all_results if not r["passed"])
    avg_score = sum(r["score"] for r in all_results) / max(len(all_results), 1)
    total_time = sum(r["elapsed"] for r in all_results)

    for r in all_results:
        status = "PASS" if r["passed"] else "FAIL"
        print(f"  Q{r['id']} {r['name']:30s} {r['score']:3d}/100 [{status}] ({r['elapsed']:.1f}s, {r['tool_calls']} calls)")

    print(f"\n  {total_pass} passed, {total_fail} failed (avg score: {avg_score:.1f}/100)")
    print(f"  Total time: {total_time:.1f}s")
    print(f"  Output: {base_output}")

    # HTML report
    generate_html_report(all_results, base_output)

    # Save summary
    with open(os.path.join(base_output, "summary.json"), "w") as f:
        json.dump({
            "results": all_results,
            "total_pass": total_pass,
            "total_fail": total_fail,
            "avg_score": avg_score,
            "total_time": total_time,
        }, f, indent=2, ensure_ascii=False)

    sys.exit(0 if total_fail == 0 else 1)


if __name__ == "__main__":
    main()
